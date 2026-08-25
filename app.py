from pathlib import Path
from io import BytesIO
from urllib.parse import urljoin, unquote
from datetime import datetime
import re

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
import urllib3
from requests.auth import HTTPBasicAuth

import base64

try:
    from PIL import Image
except ImportError:
    Image = None

import io

def crear_excel_exportacion(df, user="", password="", verify_ssl=False):
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    archivo_temporal = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    ruta_archivo = archivo_temporal.name 
    archivo_temporal.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    columnas = [
        ("COD_ITEM", "Código"),
        ("DESCRIPCION", "Producto"),
        ("MACROCATEGORIA", "Macrocategoría"),
        ("CATEGORIA", "Categoría"),
        ("SUBCATEGORIA", "Subcategoría"),
        ("FECHA_CREACION", "Fecha de creación"),
        ("TIENE_IMAGEN", "Imagen"),
    ]
    columnas = [(c, n) for c, n in columnas if c in df.columns]

    for col_idx, (_, nombre) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24

    imagen_col = next(
        (i for i, (_, nombre) in enumerate(columnas, start=1) if nombre == "Imagen"),
        None
    )

    for row_idx, (_, producto) in enumerate(df.iterrows(), start=2):
        for col_idx, (campo, nombre) in enumerate(columnas, start=1):
            valor = producto.get(campo, "")

            if campo == "FECHA_CREACION":
                if pd.notna(valor):
                    try:
                        valor = pd.to_datetime(valor).strftime("%Y-%m-%d")
                    except Exception:
                        valor = str(valor)
                else:
                    valor = ""
            elif campo == "TIENE_IMAGEN":
                valor = "Sí" if bool(valor) else "No"

            if campo != "TIENE_IMAGEN":
                if pd.isna(valor) if not isinstance(valor, (list, dict)) else False:
                    valor = ""
                ws.cell(row=row_idx, column=col_idx, value=valor)

        if imagen_col:
            cell = ws.cell(row=row_idx, column=imagen_col)
            cell.value = "Sin imagen"

            urls = producto.get("IMAGENES", [])
            if not isinstance(urls, list) or not urls:
                url = producto.get("URL_IMAGEN", "")
                urls = [url] if url else []

            if urls and user and password:
                try:
                    response = requests.get(
                        str(urls[0]),
                        auth=HTTPBasicAuth(user, password),
                        verify=verify_ssl,
                        timeout=15,
                    )
                    response.raise_for_status()

                    if Image is not None and response.content:
                        imagen = Image.open(BytesIO(response.content)).convert("RGB")
                        imagen.thumbnail((90, 90), Image.Resampling.LANCZOS)

                        buffer = BytesIO()
                        imagen.save(buffer, format="JPEG", quality=82)
                        buffer.seek(0)

                        xl_img = XLImage(buffer)
                        xl_img.width = imagen.width
                        xl_img.height = imagen.height
                        ws.add_image(xl_img, cell.coordinate)
                        cell.value = ""
                        ws.row_dimensions[row_idx].height = 72
                except Exception:
                    pass

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    thin = Side(style="thin", color="D9E1F2")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal=cell.alignment.horizontal or "left"
            )

    anchos = {
        "Código": 16,
        "Producto": 42,
        "Macrocategoría": 22,
        "Categoría": 24,
        "Subcategoría": 28,
        "Fecha de creación": 18,
        "Imagen": 16,
    }
    for col_idx, (_, nombre) in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(nombre, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    wb.save(ruta_archivo)
    return ruta_archivo

# ============================================================
# HELPER IMAGEN
# ============================================================

def bytes_a_base64_uri(contenido_bytes):
    if not contenido_bytes:
        return None
    encoded = base64.b64encode(contenido_bytes).decode("utf-8")
    return f"data:image/jpeg;base64,{encoded}"

# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="Catálogo de Productos",
    page_icon="🛍️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

BASE_DIR = Path(__file__).resolve().parent
LOGO = BASE_DIR / "assets" / "logo.png"

ARCHIVO_EXCEL = Path(
    r"C:\Users\hmbelalcazars\OneDrive - CorporacionGPF - Femsa Salud\ITEM MASTER ORIGINAL - Copia.xlsx"
)

HOJA_EXCEL = "MASTER_ITEM"

WEBDAV_SERVER = "https://staging-na01-difarma.demandware.net"
WEBDAV_ROOT = (
    "/on/demandware.servlet/webdav/"
    "Sites/Catalogs/masterCatalog_FybecaEcuador/default/images/large"
)

MAX_RESULTADOS_CATALOGO = 60
MAX_FILAS_TABLA = 500
MAX_PROPFIND_ITEMS = 50000

# ============================================================
# ESTILOS
# ============================================================

st.markdown(
    """
    <style>
    .main { padding-top: 1rem; }
    .titulo-dashboard { font-size: 32px; font-weight: 700; margin-bottom: 0; }
    .subtitulo-dashboard { font-size: 15px; color: #666; margin-bottom: 20px; }
    .linea { border-bottom: 1px solid #E5E7EB; margin: 10px 0 20px; }
    .card { background: white; padding: 20px; border-radius: 10px; border: 1px solid #E5E7EB; box-shadow: 0 2px 6px rgba(0,0,0,.05); text-align: center; }
    .card-title { font-size: 14px; color: #6B7280; }
    .card-value { font-size: 28px; font-weight: 700; }
    .producto-card { border: 1px solid #E5E7EB; border-radius: 14px; padding: 12px; margin-bottom: 10px; background: white; }
    .producto-sku { font-size: 13px; color: #6B7280; margin-top: 8px; }
    .producto-titulo { font-size: 15px; font-weight: 700; line-height: 1.25; margin-top: 4px; min-height: 40px; }
    .producto-categoria { font-size: 12px; color: #6B7280; margin-top: 6px; }
    .sin-imagen { height: 240px; display: flex; align-items: center; justify-content: center; border-radius: 10px; background: #F3F4F6; color: #6B7280; font-size: 18px; font-weight: 600; }
    [data-testid="stSidebar"] { display: none; }
    [data-testid="stSidebarCollapsedControl"] { display: none; }
    [data-testid="stToolbar"] { visibility: hidden; height: 0; }
    header { visibility: hidden; height: 0; }
    .block-container { padding-top: 1.2rem; max-width: 1500px; }
    .cliente-header { display: flex; align-items: center; justify-content: space-between; gap: 20px; margin-bottom: 8px; }
    .cliente-badge { font-size: 12px; color: #6B7280; background: #F3F4F6; border-radius: 999px; padding: 6px 12px; white-space: nowrap; }
    .section-title { font-size: 20px; font-weight: 700; margin: 22px 0 10px; }
    @media print {
        @page { size: A4 landscape; margin: 10mm; }
        body { background: white !important; }
        [data-testid="stSidebar"], [data-testid="stSidebarCollapsedControl"],
        [data-testid="stToolbar"], header, footer, button,
        [data-testid="stDownloadButton"], [data-testid="stRadio"] { display: none !important; }
        .block-container { max-width: 100% !important; padding: 0 !important; }
        .card { box-shadow: none !important; break-inside: avoid; }
        .producto-card { break-inside: avoid; box-shadow: none !important; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# CREDENCIALES
# ============================================================

def obtener_credenciales():
    try:
        cfg = st.secrets["webdav"]
        user = str(cfg["user"])
        password = str(cfg["password"])
        verify_ssl = cfg.get("verify_ssl", False)
        if isinstance(verify_ssl, str):
            verify_ssl = verify_ssl.lower() in ("true", "1", "yes", "si")
        return user, password, bool(verify_ssl), None
    except Exception as e:
        return "", "", False, (
            "No se encontraron las credenciales WebDAV. "
            "Crea .streamlit/secrets.toml con [webdav], user y password. "
            f"Detalle: {type(e).__name__}: {e}"
        )

WEBDAV_USER, WEBDAV_PASSWORD, VERIFY_SSL, ERROR_CREDENCIALES = obtener_credenciales()

if not VERIFY_SSL:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================
# MASTER ITEM
# ============================================================

@st.cache_data(show_spinner="Actualizando información...")
def cargar_master_item(ruta):
    df = pd.read_excel(
        ruta,
        sheet_name=HOJA_EXCEL,
        engine="openpyxl",
    )
    
    df.columns = df.columns.astype(str).str.strip()

    for col in ["COD_ITEM", "DESCRIPCION", "MACROCATEGORIA", "CATEGORIA", "SUBCATEGORIA"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    if "FECHA_CREACION" in df.columns:
        df["FECHA_CREACION"] = pd.to_datetime(df["FECHA_CREACION"], errors="coerce")
        df["ANIO"] = df["FECHA_CREACION"].dt.year
        df["MES"] = df["FECHA_CREACION"].dt.month
        df["DIA"] = df["FECHA_CREACION"].dt.day
    else:
        df["ANIO"] = pd.Series(pd.NA, index=df.index, dtype="Int64")
        df["MES"] = pd.Series(pd.NA, index=df.index, dtype="Int64")
        df["DIA"] = pd.Series(pd.NA, index=df.index, dtype="Int64")

    return df

def normalizar_sku(valor):
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        texto = texto[:-2]
    return texto

# ============================================================
# PROPFIND RECURSIVO
# ============================================================

def extraer_hrefs(xml_text):
    from xml.etree import ElementTree as ET
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []

    hrefs = []
    for elem in root.iter():
        if elem.tag.lower().endswith("href") and elem.text:
            hrefs.append(elem.text.strip())
    return hrefs

def obtener_nombre_archivo(href):
    limpio = unquote(str(href).split("?")[0].split("#")[0])
    return limpio.rstrip("/").split("/")[-1]

def obtener_sku_desde_nombre(nombre):
    nombre = unquote(str(nombre)).strip()
    match = re.match(r"^(\d+)", nombre)
    if match:
        return match.group(1)
    return ""

@st.cache_data(ttl=1800, show_spinner="Preparando catálogo de imágenes...")
def construir_indice_webdav(server, root, user, password, verify_ssl):
    session = requests.Session()
    session.auth = HTTPBasicAuth(user, password)
    session.verify = verify_ssl
    session.headers.update({"User-Agent": "Fybeca-Dashboard-Catalogo/3.0"})

    indice = {}
    carpetas_pendientes = [root.rstrip("/")]
    carpetas_visitadas = set()
    archivos_procesados = 0

    while carpetas_pendientes:
        carpeta = carpetas_pendientes.pop()
        if carpeta in carpetas_visitadas:
            continue

        carpetas_visitadas.add(carpeta)
        url = server.rstrip("/") + carpeta

        try:
            response = session.request(
                "PROPFIND",
                url,
                headers={"Depth": "1", "Content-Type": "application/xml"},
                timeout=30,
            )
            response.raise_for_status()
        except Exception:
            continue

        hrefs = extraer_hrefs(response.text)

        for href in hrefs:
            href_decoded = unquote(href)

            if href_decoded.startswith("http://") or href_decoded.startswith("https://"):
                url_item = href_decoded
            else:
                url_item = urljoin(server.rstrip("/") + "/", href_decoded.lstrip("/"))

            path = href_decoded.rstrip("/")
            nombre = obtener_nombre_archivo(path)

            if not nombre:
                continue

            if href_decoded.endswith("/"):
                siguiente = path[len(server):] if path.startswith(server) else "/" + path.lstrip("/")
                if siguiente not in carpetas_visitadas:
                    carpetas_pendientes.append(siguiente)
                continue

            extension = Path(nombre).suffix.lower()
            if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
                continue

            sku = obtener_sku_desde_nombre(nombre)
            if not sku:
                continue

            archivos_procesados += 1
            indice.setdefault(sku, [])
            if url_item not in indice[sku]:
                indice[sku].append(url_item)

            if archivos_procesados >= MAX_PROPFIND_ITEMS:
                return indice

    return indice

# ============================================================
# CARGA UNIFICADA DE DATOS
# ============================================================

error_excel = None
df = pd.DataFrame()

try:
    if not ARCHIVO_EXCEL.exists():
        error_excel = "No fue posible cargar la información del Excel."
    else:
        df = cargar_master_item(str(ARCHIVO_EXCEL))
except Exception as e:
    error_excel = f"Error al cargar Excel: {e}"

if not df.empty and not ERROR_CREDENCIALES:
    try:
        indice_imagenes = construir_indice_webdav(
            WEBDAV_SERVER,
            WEBDAV_ROOT,
            WEBDAV_USER,
            WEBDAV_PASSWORD,
            VERIFY_SSL,
        )

        df["IMAGENES"] = df["COD_ITEM"].apply(
            lambda sku: indice_imagenes.get(normalizar_sku(sku), [])
        )
        df["URL_IMAGEN"] = df["IMAGENES"].apply(
            lambda urls: urls[0] if isinstance(urls, list) and urls else ""
        )
        df["TIENE_IMAGEN"] = df["IMAGENES"].apply(
            lambda urls: isinstance(urls, list) and len(urls) > 0
        )
    except Exception as e:
        st.warning(f"No fue posible consultar imágenes WebDAV: {e}")
        df["IMAGENES"] = [[] for _ in range(len(df))]
        df["URL_IMAGEN"] = ""
        df["TIENE_IMAGEN"] = False
else:
    df["IMAGENES"] = [[] for _ in range(len(df))]
    df["URL_IMAGEN"] = ""
    df["TIENE_IMAGEN"] = False


# ============================================================
# DESCARGAR IMAGEN
# ============================================================

@st.cache_data(max_entries=500, show_spinner=False)
def descargar_imagen(url, user, password, verify_ssl):
    if not url:
        return None
    try:
        response = requests.get(
            url,
            auth=HTTPBasicAuth(user, password),
            verify=verify_ssl,
            timeout=20,
        )
        response.raise_for_status()

        if not response.content:
            return None

        if Image is None:
            return response.content

        imagen = Image.open(BytesIO(response.content))
        if imagen.mode not in ("RGB", "L"):
            imagen = imagen.convert("RGB")

        imagen.thumbnail((420, 420), Image.Resampling.LANCZOS)
        salida = BytesIO()
        imagen.save(salida, format="JPEG", quality=82, optimize=True)
        return salida.getvalue()
    except Exception:
        return None

# ============================================================
# INTERFAZ Y FILTROS
# ============================================================

st.markdown('<div class="section-title">🔎 Buscar producto</div>', unsafe_allow_html=True)

buscar = st.text_input(
    "Buscar producto",
    placeholder="Ingrese SKU o nombre del producto...",
    key="buscar_sku",
)

st.markdown('<div class="section-title">🎯 Filtrar catálogo</div>', unsafe_allow_html=True)

def opciones_columna(df_base, columna):
    if columna not in df_base.columns:
        return ["Todas"]
    valores = df_base[columna].fillna("").astype(str).str.strip()
    valores = [x for x in valores.unique() if x]
    return ["Todas"] + sorted(valores)

macro_opciones = opciones_columna(df, "MACROCATEGORIA")
categoria_opciones = opciones_columna(df, "CATEGORIA")
subcategoria_opciones = opciones_columna(df, "SUBCATEGORIA")

st.write("Shape:", df.shape)
st.write("Columnas:", df.columns.tolist())
st.dataframe(df.head())

anios = pd.to_numeric(df["ANIO"], errors="coerce").dropna().astype(int).unique().tolist()
anio_opciones = ["Todos"] + sorted(anios, reverse=True)

MESES = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5, "Junio": 6,
    "Julio": 7, "Agosto": 8, "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12,
}

col1, col2, col3 = st.columns(3)
with col1:
    macro = st.selectbox("Macrocategoría", macro_opciones, key="filtro_macrocategoria")
with col2:
    categoria = st.selectbox("Categoría", categoria_opciones, key="filtro_categoria")
with col3:
    subcategoria = st.selectbox("Subcategoría", subcategoria_opciones, key="filtro_subcategoria")

col4, col5, col6 = st.columns(3)
with col4:
    anio = st.selectbox("Año", anio_opciones, key="filtro_anio")
with col5:
    mes = st.selectbox("Mes", ["Todos"] + list(MESES.keys()), key="filtro_mes")
with col6:
    dia = st.selectbox("Día", ["Todos"] + list(range(1, 32)), key="filtro_dia")

filtro_imagen = st.selectbox(
    "🖼️ Disponibilidad de imagen",
    ["Todos", "Con imagen", "Sin imagen"],
    key="filtro_imagen",
)

# ============================================================
# APLICAR FILTROS
# ============================================================

df_filtrado = df.copy()

if buscar:
    texto = buscar.strip().lower()
    mascara = pd.Series(False, index=df_filtrado.index)
    mascara |= df_filtrado["COD_ITEM"].astype(str).str.lower().str.contains(texto, na=False, regex=False)

    if "DESCRIPCION" in df_filtrado.columns:
        mascara |= df_filtrado["DESCRIPCION"].astype(str).str.lower().str.contains(texto, na=False, regex=False)

    df_filtrado = df_filtrado[mascara]

if macro != "Todas":
    df_filtrado = df_filtrado[df_filtrado["MACROCATEGORIA"] == macro]

if categoria != "Todas":
    df_filtrado = df_filtrado[df_filtrado["CATEGORIA"] == categoria]

if subcategoria != "Todas":
    df_filtrado = df_filtrado[df_filtrado["SUBCATEGORIA"] == subcategoria]

if anio != "Todos":
    df_filtrado = df_filtrado[pd.to_numeric(df_filtrado["ANIO"], errors="coerce") == int(anio)]

if mes != "Todos":
    df_filtrado = df_filtrado[pd.to_numeric(df_filtrado["MES"], errors="coerce") == MESES[mes]]

if dia != "Todos":
    df_filtrado = df_filtrado[pd.to_numeric(df_filtrado["DIA"], errors="coerce") == int(dia)]

if filtro_imagen == "Con imagen":
    df_filtrado = df_filtrado[df_filtrado["TIENE_IMAGEN"] == True]
elif filtro_imagen == "Sin imagen":
    df_filtrado = df_filtrado[df_filtrado["TIENE_IMAGEN"] == False]

# ============================================================
# DEDUPLICAR PRODUCTOS (SUELTA sobre CAJA cuando tienen distinto SKU)
# ============================================================
if not df_filtrado.empty and "DESCRIPCION" in df_filtrado.columns:
    # 1. Función para crear un nombre base quitando las palabras CAJA y SUELTA
    def obtener_nombre_base(desc):
        texto = str(desc).upper()
        # Remueve las palabras CAJA o SUELTA (y espacios extra sobrantes)
        texto_limpio = re.sub(r'\b(CAJA|SUELTA)\b', '', texto)
        return " ".join(texto_limpio.split())

    # 2. Función para definir prioridad (1: SUELTA, 2: CAJA, 3: OTROS)
    def tipo_empaque(desc):
        texto = str(desc).upper()
        if "SUELTA" in texto:
            return 1
        elif "CAJA" in texto:
            return 2
        return 3

    # 3. Crear columnas auxiliares
    df_filtrado["_NOMBRE_BASE"] = df_filtrado["DESCRIPCION"].apply(obtener_nombre_base)
    df_filtrado["_PRIORIDAD_EMPAQUE"] = df_filtrado["DESCRIPCION"].apply(tipo_empaque)

    # 4. Ordenar por el Nombre Base y por la prioridad asignada
    df_filtrado = df_filtrado.sort_values(by=["_NOMBRE_BASE", "_PRIORIDAD_EMPAQUE"])

    # 5. Eliminar duplicados basándose en el Nombre Base (conserva la versión SUELTA)
    df_filtrado = df_filtrado.drop_duplicates(subset=["_NOMBRE_BASE"], keep="first")

    # 6. Eliminar columnas auxiliares
    df_filtrado = df_filtrado.drop(columns=["_NOMBRE_BASE", "_PRIORIDAD_EMPAQUE"])

# ============================================================
# RESULTADOS
# ============================================================

st.markdown("---")
st.markdown(f'<div class="section-title">📋 Resultados ({len(df_filtrado):,} productos)</div>', unsafe_allow_html=True)

vista = st.radio("Vista", ["🖼️ Catálogo", "📋 Tabla"], horizontal=True)

# ============================================================
# VISTA CATÁLOGO
# ============================================================

if vista == "🖼️ Catálogo":
    if df_filtrado.empty:
        st.warning("No existen productos con los filtros seleccionados.")
    else:
        if len(df_filtrado) > MAX_RESULTADOS_CATALOGO:
            st.info(
                f"Se encontraron {len(df_filtrado):,} productos. "
                f"Se muestran los primeros {MAX_RESULTADOS_CATALOGO:,}."
            )

        datos = df_filtrado.head(MAX_RESULTADOS_CATALOGO)

        for inicio in range(0, len(datos), 4):
            fila = datos.iloc[inicio:inicio + 4]
            columnas = st.columns(4)

            for pos, (_, producto) in enumerate(fila.iterrows()):
                with columnas[pos]:
                    sku = normalizar_sku(producto["COD_ITEM"])
                    descripcion = str(producto.get("DESCRIPCION", ""))
                    categoria_producto = str(producto.get("CATEGORIA", ""))
                    subcategoria_producto = str(producto.get("SUBCATEGORIA", ""))

                    st.markdown('<div class="producto-card">', unsafe_allow_html=True)

                    imagenes_producto = producto.get("IMAGENES", [])
                    if not isinstance(imagenes_producto, list):
                        imagenes_producto = []

                    if imagenes_producto:
                        imagen_principal = descargar_imagen(
                            imagenes_producto[0],
                            WEBDAV_USER,
                            WEBDAV_PASSWORD,
                            VERIFY_SSL,
                        )

                        if imagen_principal:
                            st.image(imagen_principal, use_container_width=True)
                        else:
                            st.markdown('<div class="sin-imagen">📷 Error al cargar imagen</div>', unsafe_allow_html=True)

                        if len(imagenes_producto) > 1:
                            with st.expander(f"🖼️ Ver galería ({len(imagenes_producto)} imágenes)"):
                                for url_extra in imagenes_producto[:4]:
                                    imagen_extra = descargar_imagen(
                                        url_extra,
                                        WEBDAV_USER,
                                        WEBDAV_PASSWORD,
                                        VERIFY_SSL,
                                    )
                                    if imagen_extra:
                                        st.image(imagen_extra, use_container_width=True)

                        st.link_button("🔗 Ver imagen", imagenes_producto[0], use_container_width=True)
                    else:
                        st.markdown('<div class="sin-imagen">📷 Sin imagen</div>', unsafe_allow_html=True)

                    st.markdown(
                        f"""
                        <div class="producto-sku">Código: {sku}</div>
                        <div class="producto-titulo">{descripcion}</div>
                        <div class="producto-categoria">
                            {categoria_producto} {" · " + subcategoria_producto if subcategoria_producto else ""}
                        </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

# ============================================================
# VISTA TABLA
# ============================================================

else:
    columnas_tabla = [
        "COD_ITEM", "DESCRIPCION", "MACROCATEGORIA",
        "CATEGORIA", "SUBCATEGORIA", "FECHA_CREACION",
        "TIENE_IMAGEN", "IMAGENES", "URL_IMAGEN"
    ]
    columnas_tabla = [c for c in columnas_tabla if c in df_filtrado.columns]

    tabla = df_filtrado[columnas_tabla].copy()

    if len(tabla) > MAX_FILAS_TABLA:
        st.info(f"Se muestran los primeros {MAX_FILAS_TABLA:,} registros.")
        tabla = tabla.head(MAX_FILAS_TABLA)

    def obtener_url_imagen(row):
        imagenes = row.get("IMAGENES", [])
        if isinstance(imagenes, list) and len(imagenes) > 0:
            return imagenes[0]
        url = row.get("URL_IMAGEN", "")
        return str(url).strip() if pd.notna(url) and str(url).strip() else None

    tabla["_URL_IMAGEN"] = tabla.apply(obtener_url_imagen, axis=1)

    # Obtenemos la miniatura llamando a descargar_imagen directamente (retorna bytes)
    def obtener_miniatura_base64(url):
        img_bytes = descargar_imagen(url, WEBDAV_USER, WEBDAV_PASSWORD, VERIFY_SSL)
        if img_bytes:
            return bytes_a_base64_uri(img_bytes)
        return None

    tabla["Imagen"] = tabla["_URL_IMAGEN"].apply(obtener_miniatura_base64)

    tabla = tabla.rename(columns={
        "COD_ITEM": "Código",
        "DESCRIPCION": "Producto",
        "MACROCATEGORIA": "Macrocategoría",
        "CATEGORIA": "Categoría",
        "SUBCATEGORIA": "Subcategoría",
        "FECHA_CREACION": "Fecha de creación",
        "TIENE_IMAGEN": "Tiene imagen",
    })

    if "Fecha de creación" in tabla.columns:
        tabla["Fecha de creación"] = pd.to_datetime(
            tabla["Fecha de creación"], errors="coerce"
        ).dt.strftime("%Y-%m-%d")

    if "Tiene imagen" in tabla.columns:
        tabla["Tiene imagen"] = tabla["Tiene imagen"].map({True: "🟢 Sí", False: "🔴 No"})

    for col_elim in ["_URL_IMAGEN", "IMAGENES", "URL_IMAGEN"]:
        if col_elim in tabla.columns:
            tabla = tabla.drop(columns=[col_elim])

    st.dataframe(
        tabla,
        use_container_width=True,
        height=550,
        hide_index=True,
        column_config={
            "Código": st.column_config.TextColumn("Código", width="small"),
            "Producto": st.column_config.TextColumn("Producto", width="large"),
            "Tiene imagen": st.column_config.TextColumn("Imagen disponible", width="small"),
            "Imagen": st.column_config.ImageColumn("Miniatura", help="Imagen del producto", width="small"),
        },
    )

# ============================================================
# IMPRESIÓN DEL DASHBOARD
# ============================================================

st.markdown("---")
st.caption("Catálogo de Productos · Reporte para consulta y revisión")
